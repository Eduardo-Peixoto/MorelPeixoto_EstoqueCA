import tkinter as tk
from tkinter import ttk, PhotoImage
from tkinter import messagebox
from PIL import Image, ImageTk
import openpyxl
import os
from datetime import datetime

# variável global para tela de login (para ficar com uma tela só)
tela_login = None

# cores
cor_preta = "#f0f3f5" 
cor_branca = "#f3ffff"
cor_verde = "#3fb5a3" 
valr = "#38576b" 
letr = "#403d3d" 

# identificação do diretório do código .py (a planilha deve estar no mesmo local, assim como as imagens!)
diretorio_atual = os.path.dirname(os.path.abspath(__file__))

# caminho para imagens
caminho_imagem_eb = os.path.join(diretorio_atual, "simbolo_eb.png")
caminho_imagem_ime = os.path.join(diretorio_atual, "simbolo_ime.png")

# caminho da planilha de estoque
ARQUIVO_ESTOQUE = os.path.join(diretorio_atual, "estoque.xlsx")

# caminho para a planilha de inventário
ARQUIVO_INVENTARIO = os.path.join(diretorio_atual, "inventario_usuarios.xlsx")

# gerenciar estoque e inventário dos usuários no excel
class Estoque:
    def __init__(self, arquivo):
        self.arquivo = arquivo
        
        try:
            # planilha de estoque
            self.wb = openpyxl.load_workbook(arquivo)
            self.sheet = self.wb.active
            
            # planilha de inventario
            self.wb_inventario = openpyxl.load_workbook(ARQUIVO_INVENTARIO)
            self.sheet_inventario = self.wb_inventario.active
        except FileNotFoundError:
            print(f"Erro: O arquivo {arquivo} ou {ARQUIVO_INVENTARIO} não foi encontrado!")
            raise

    # função para salvar as planilhas        
    def salvar(self):
        self.wb.save(self.arquivo)

    def exibir_estoque(self):
        return [(row[0].value, row[1].value) for row in self.sheet.iter_rows(min_row=2, max_col=2)]

    def encontrar_item(self, item_nome):
        for row in self.sheet.iter_rows(min_row=2, max_col=3):
            if row[0].value == item_nome:
                return row
        return None

    def retirar_item(self, item_nome, quantidade, usuario):
        linha_item = self.encontrar_item(item_nome)
        if linha_item:
            quantidade_atual = linha_item[1].value
            if quantidade_atual >= quantidade:
                linha_item[1].value -= quantidade
                self.salvar()
                return True
        return False

    def devolver_item(self, item_nome, quantidade, usuario):
        linha_item = self.encontrar_item(item_nome)
        if linha_item:
            linha_item[1].value += quantidade
            self.salvar()
            return True
        return False
    
    def atualizar_inventario_usuario(self, nome_usuario, item_nome, quantidade, acao):
        # encontrar o índice do usuário na planilha
        for row in self.sheet_inventario.iter_rows(min_row=2, max_col=13):
            if row[0].value == nome_usuario:
                # encontra a coluna do item
                for idx, cell in enumerate(row[1:], start=1):  
                    if self.sheet.cell(row=idx+1, column=1).value == item_nome:  
                        if acao == "cautelar":
                            # atualiza a quantidade para o item cautelado
                            row[idx].value = (row[idx].value or 0) + quantidade
                        elif acao == "devolver":
                            # verifica se o usuário tem a quantidade suficiente para devolver
                            if row[idx].value >= quantidade:
                                row[idx].value -= quantidade
                            else:
                                return False  # quantidade insuficiente no inventário
                self.wb_inventario.save(ARQUIVO_INVENTARIO)  # salvar depois de alterar
                return True  
        return False  
    
    def adicionar_item(self, item_nome, quantidade):
        linha_item = self.encontrar_item(item_nome)
        if linha_item:
            # Se o item já existe, adicionar a quantidade ao existente
            linha_item[1].value += quantidade
        else:
            # Se o item não existe, criar uma nova linha no estoque
            nova_linha = self.sheet.max_row + 1
            self.sheet.cell(row=nova_linha, column=1).value = item_nome
            self.sheet.cell(row=nova_linha, column=2).value = quantidade

            # Atualizar a planilha de inventário
            # Adicionar uma nova coluna para o item no inventário
            coluna_item = self.sheet_inventario.max_column + 1
            self.sheet_inventario.cell(row=1, column=coluna_item).value = item_nome  # Cabeçalho do novo item
            for row in self.sheet_inventario.iter_rows(min_row=2, max_col=coluna_item):
                # Inicializar a quantidade do novo item como 0 para todos os usuários
                row[coluna_item - 1].value = 0

        # Salvar as atualizações nos arquivos
        self.salvar()
        self.wb_inventario.save(ARQUIVO_INVENTARIO)
    
        return True


    

# funções de integração com a interface gráfica
estoque = Estoque(ARQUIVO_ESTOQUE)
usuario_logado = None

usuarios_info = {
    "morel@ime.eb.br": {"nome": "Morel"},
    "peixoto@ime.eb.br": {"nome": "Peixoto"},
    "nicolas@ime.eb.br": {"nome": "Nicolas"},
}

def realizar_login(email, senha):
    global usuario_logado
    
    # usuários para autenticação
    usuarios = {"morel@ime.eb.br": "morel", "peixoto@ime.eb.br": "peixoto", "nicolas@ime.eb.br": "nicolas", "encmat@ime.eb.br": "encmat"}

    if email in usuarios and usuarios[email] == senha:
        usuario_logado = email
        return True
    return False

def login_admin(email, senha):
    if realizar_login(email, senha):
        for widget in tela_login.winfo_children():
            widget.destroy()
        if email == "encmat@ime.eb.br":
            criar_tela_adicao_itens(tela_login)  # interface exclusiva para encmat
        else:
            criar_tela_estoque(tela_login)  # interface padrão para outros usuários
    else:
        messagebox.showerror("Erro", "Usuário ou senha inválidos.")

def cautelar_item(item, quantidade):
    global usuario_logado
    if item and quantidade.isdigit() and int(quantidade) > 0:
        if usuario_logado in usuarios_info:
            info_usuario = usuarios_info[usuario_logado]
            nome_responsavel = info_usuario["nome"]
                
            if estoque.retirar_item(item, int(quantidade), nome_responsavel):
                if estoque.atualizar_inventario_usuario(nome_responsavel, item, int(quantidade), "cautelar"):
                    messagebox.showinfo("Sucesso", f"{quantidade} unidades de {item} cauteladas por {nome_responsavel}.")
                else:
                    messagebox.showerror("Erro", "Falha ao atualizar o inventário do usuário.")
            else:
                messagebox.showerror("Erro", "Estoque insuficiente ou item não encontrado.")
        else:
            messagebox.showerror("Erro", "Usuário não registrado para retirada.")
    else:
        messagebox.showerror("Erro", "Selecione um item válido e insira uma quantidade.")



def devolver_item(item, quantidade):
    global usuario_logado
    if item and quantidade.isdigit() and int(quantidade) > 0:
        if usuario_logado in usuarios_info:
            info_usuario = usuarios_info[usuario_logado]
            nome_responsavel = info_usuario["nome"]

            if estoque.atualizar_inventario_usuario(nome_responsavel, item, int(quantidade), "devolver"):
                if estoque.devolver_item(item, int(quantidade), nome_responsavel):
                    messagebox.showinfo("Sucesso", f"{quantidade} unidades de {item} devolvidas.")
                else:
                    messagebox.showerror("Erro", "Item não encontrado no estoque.")
            else:
                messagebox.showerror("Erro", "Quantidade insuficiente no inventário do usuário.")
        else:
            messagebox.showerror("Erro", "Usuário não registrado.")
    else:
        messagebox.showerror("Erro", "Selecione um item válido e insira uma quantidade.")

def adicionar_item_estoque(item_nome, quantidade):
    if item_nome and quantidade.isdigit() and int(quantidade) > 0:
        if estoque.adicionar_item(item_nome, int(quantidade)):
            messagebox.showinfo("Sucesso", f"{quantidade} unidades de {item_nome} adicionadas ao estoque.")
        else:
            messagebox.showerror("Erro", "Falha ao adicionar item ao estoque.")
    else:
        messagebox.showerror("Erro", "Nome do item e quantidade válidos são necessários.")


# interface gráfica
def criar_tela_login():
    global tela_login
    tela_login = tk.Tk()
    tela_login.title("Login")
    tela_login.geometry("600x325")
    tela_login.configure(background=cor_verde)

    tk.Label(tela_login, text="Login", font=('Arial', 18), bg=cor_verde, fg=letr).pack(pady=20)

    img1 = Image.open(caminho_imagem_eb)
    img1 = img1.resize((100, 120), Image.Resampling.LANCZOS) 
    imagem1 = ImageTk.PhotoImage(img1) 

    img2 = Image.open(caminho_imagem_ime)
    img2 = img2.resize((100, 120), Image.Resampling.LANCZOS) 
    imagem2 = ImageTk.PhotoImage(img2)

    img_label1 = tk.Label(tela_login, image=imagem1, bg=cor_verde)
    img_label1.place(x=0, y=0)

    img_label2 = tk.Label(tela_login, image=imagem2, bg=cor_verde)
    img_label2.place(x=500, y=0)

    tk.Label(tela_login, text="E-Mail:", bg=cor_verde, fg=letr).pack()
    entrada_email = tk.Entry(tela_login)
    entrada_email.pack(pady=5)

    tk.Label(tela_login, text="Senha:", bg=cor_verde, fg=letr).pack()
    entrada_senha = tk.Entry(tela_login, show="*")
    entrada_senha.pack(pady=5)

    botao_login = ttk.Button(tela_login, text="Login", command=lambda: login_admin(entrada_email.get(), entrada_senha.get()))
    botao_login.pack(pady=20)

    tela_login.mainloop()

def exibir_estoque_em_tela():

    tela_itens = tk.Toplevel()
    tela_itens.title("Itens no Estoque")
    tela_itens.geometry("500x400")
    tela_itens.configure(background=cor_branca)

    # nova tela
    tk.Label(tela_itens, text="Itens Disponíveis no Estoque", font=('Arial', 16), bg=cor_branca, fg=valr).pack(pady=10)

    # lista de itens
    frame_itens = tk.Frame(tela_itens, bg=cor_branca)
    frame_itens.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    # exibir itens do estoque
    itens = estoque.exibir_estoque()
    for item, quantidade in itens:
        tk.Label(frame_itens, text=f"{item}: {quantidade}", font=('Arial', 12), bg=cor_branca, fg=letr).pack(anchor="w", pady=2)

def criar_tela_estoque(tela):
    tela.title("Reserva de Materiais")
    tela.geometry("500x400")
    tela.configure(background=cor_verde)

    tk.Label(tela, text="Estoque de Materiais", font=('Arial', 18), bg=cor_verde, fg=letr).pack(pady=20)

    itens = estoque.exibir_estoque()
    item_selecionado = tk.StringVar(tela)
    ttk.Combobox(tela, textvariable=item_selecionado, values=[i[0] for i in itens]).pack(pady=10)

    tk.Label(tela, text="Quantidade:", bg=cor_verde, fg=letr).pack()
    entrada_quantidade = tk.Entry(tela)
    entrada_quantidade.pack(pady=5)

    ttk.Button(tela, text="Cautelar", command=lambda: cautelar_item(item_selecionado.get(), entrada_quantidade.get())).pack(pady=10)
    ttk.Button(tela, text="Devolver", command=lambda: devolver_item(item_selecionado.get(), entrada_quantidade.get())).pack(pady=10)
    ttk.Button(tela, text="Mostrar Estoque", command=exibir_estoque_em_tela).pack(pady=10)

def criar_tela_adicao_itens(tela):
    tela.title("Adicionar Itens ao Estoque")
    tela.geometry("400x300")
    tela.configure(background=cor_verde)

    tk.Label(tela, text="Adicionar Itens ao Estoque", font=('Arial', 18), bg=cor_verde, fg=letr).pack(pady=20)

    tk.Label(tela, text="Nome do Item:", bg=cor_verde, fg=letr).pack()
    entrada_item_nome = tk.Entry(tela)
    entrada_item_nome.pack(pady=5)

    tk.Label(tela, text="Quantidade:", bg=cor_verde, fg=letr).pack()
    entrada_quantidade = tk.Entry(tela)
    entrada_quantidade.pack(pady=5)

    ttk.Button(
        tela,
        text="Adicionar",
        command=lambda: adicionar_item_estoque(entrada_item_nome.get(), entrada_quantidade.get())
    ).pack(pady=20)


# iniciar a aplicação
criar_tela_login()