import tkinter as tk
from tkinter import ttk, PhotoImage
from tkinter import messagebox
from PIL import Image, ImageTk
import openpyxl
from datetime import datetime

# Paleta de cores
cor_preta = "#f0f3f5" 
cor_branca = "#f3ffff"
cor_verde = "#3fb5a3" 
valr = "#38576b" 
letr = "#403d3d" 

# Arquivo de estoque
ARQUIVO_ESTOQUE = "estoque.xlsx"

# Classe para gerenciar estoque no Excel
class Estoque:
    def __init__(self, arquivo):
        self.arquivo = arquivo
        self.wb = openpyxl.load_workbook(arquivo)
        self.sheet = self.wb.active

    def salvar(self):
        self.wb.save(self.arquivo)

    def exibir_estoque(self):
        return [(row[0].value, row[1].value) for row in self.sheet.iter_rows(min_row=2, max_col=2)]

    def encontrar_item(self, item_nome):
        for row in self.sheet.iter_rows(min_row=2, max_col=3):
            if row[0].value == item_nome:
                return row
        return None

    def retirar_item(self, item_nome, quantidade, usuario):
        linha_item = self.encontrar_item(item_nome)
        if linha_item:
            quantidade_atual = linha_item[1].value
            if quantidade_atual >= quantidade:
                linha_item[1].value -= quantidade
                historico = linha_item[2].value or ""
                linha_item[2].value = f"{historico}\n{datetime.now()} - {usuario} retirou {quantidade} "
                self.salvar()
                return True
        return False

    def devolver_item(self, item_nome, quantidade, usuario):
        linha_item = self.encontrar_item(item_nome)
        if linha_item:
            linha_item[1].value += quantidade
            historico = linha_item[2].value or ""
            linha_item[2].value = f"{historico}\n{datetime.now()} - {usuario} devolveu {quantidade}"
            self.salvar()
            return True
        return False

# Funções de integração com a interface
estoque = Estoque(ARQUIVO_ESTOQUE)
usuario_logado = None

def realizar_login(email, senha):
    global usuario_logado
    # Usuários para autenticação (substituir por banco de dados em um sistema real)
    usuarios = {"admin@example.com": "admin123", "user@example.com": "user123", "1":"2"}

    if email in usuarios and usuarios[email] == senha:
        usuario_logado = email
        return True
    return False

def login_admin(email, senha):
    if realizar_login(email, senha):
        criar_tela_estoque()
    else:
        messagebox.showerror("Erro", "Usuário ou senha inválidos.")

def cautelar_item(item, quantidade):
    global usuario_logado
    if estoque.retirar_item(item, int(quantidade), usuario_logado):
        messagebox.showinfo("Sucesso", f"{quantidade} unidades de {item} cauteladas.")
    else:
        messagebox.showerror("Erro", "Estoque insuficiente ou item não encontrado.")

def devolver_item(item, quantidade):
    global usuario_logado
    if estoque.devolver_item(item, int(quantidade), usuario_logado):
        messagebox.showinfo("Sucesso", f"{quantidade} unidades de {item} devolvidas.")
    else:
        messagebox.showerror("Erro", "Item não encontrado.")

# Telas da interface gráfica
def criar_tela_login():
    tela_login = tk.Tk()
    tela_login.title("Login")
    tela_login.geometry("600x325")
    tela_login.configure(background=cor_verde)

    tk.Label(tela_login, text="Login", font=('Arial', 18), bg=cor_verde, fg=letr).pack(pady=20)

    img1 = Image.open("C:/Users/amado/Documents/VSCodeProjects/ProgAp/simbolo_eb.png")
    img1 = img1.resize((100, 120), Image.Resampling.LANCZOS) 
    imagem1 = ImageTk.PhotoImage(img1) 
   
    img2 = Image.open("C:/Users/amado/Documents/VSCodeProjects/ProgAp/simbolo_ime.png")
    img2 = img2.resize((100, 120), Image.Resampling.LANCZOS) 
    imagem2 = ImageTk.PhotoImage(img2)

    img_label1 = tk.Label(tela_login, image=imagem1, bg=cor_verde)
    img_label1.place(x=0, y=0)

    img_label2 = tk.Label(tela_login, image=imagem2, bg=cor_verde)
    img_label2.place(x=500, y=0)


    tk.Label(tela_login, text="E-Mail:", bg=cor_verde, fg=letr).pack()
    entrada_email = tk.Entry(tela_login)
    entrada_email.pack(pady=5)

    tk.Label(tela_login, text="Senha:", bg=cor_verde, fg=letr).pack()
    entrada_senha = tk.Entry(tela_login, show="*")
    entrada_senha.pack(pady=5)

    botao_login = ttk.Button(tela_login, text="Login", command=lambda: login_admin(entrada_email.get(), entrada_senha.get()))
    botao_login.pack(pady=20)

    tela_login.mainloop()

def exibir_estoque_em_tela():
    #Exibe os itens do estoque em uma nova janela.
    # Criar nova janela

    tela_itens = tk.Toplevel()
    tela_itens.title("Itens no Estoque")
    tela_itens.geometry("500x400")
    tela_itens.configure(background=cor_branca)

    # Título da nova tela
    tk.Label(tela_itens, text="Itens Disponíveis no Estoque", font=('Arial', 16), bg=cor_branca, fg=valr).pack(pady=10)

    # Frame para a lista de itens
    frame_itens = tk.Frame(tela_itens, bg=cor_branca)
    frame_itens.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    # Exibir os itens do estoque
    itens = estoque.exibir_estoque()
    for item, quantidade in itens:
        tk.Label(frame_itens, text=f"{item}: {quantidade}", font=('Arial', 12), bg=cor_branca, fg=letr).pack(anchor="w", pady=2)

def criar_tela_estoque():
    tela_estoque = tk.Toplevel()
    tela_estoque.title("Reserva de Materiais")
    tela_estoque.geometry("500x400")
    tela_estoque.configure(background=cor_verde)

    tk.Label(tela_estoque, text="Estoque de Materiais", font=('Arial', 18), bg=cor_verde, fg=letr).pack(pady=20)

    itens = estoque.exibir_estoque()
    item_selecionado = tk.StringVar(tela_estoque)
    ttk.Combobox(tela_estoque, textvariable=item_selecionado, values=[i[0] for i in itens]).pack(pady=10)

    tk.Label(tela_estoque, text="Quantidade:", bg=cor_verde, fg=letr).pack()
    entrada_quantidade = tk.Entry(tela_estoque)
    entrada_quantidade.pack(pady=5)

    ttk.Button(tela_estoque, text="Cautelar", command=lambda: cautelar_item(item_selecionado.get(), entrada_quantidade.get())).pack(pady=10)
    ttk.Button(tela_estoque, text="Devolver", command=lambda: devolver_item(item_selecionado.get(), entrada_quantidade.get())).pack(pady=10)
    ttk.Button(tela_estoque, text="Mostrar Estoque", command=exibir_estoque_em_tela).pack(pady=10)

# Inicializa a aplicação
criar_tela_login()